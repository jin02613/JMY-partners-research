# -*- coding: utf-8 -*-
"""
공유 유틸리티 모듈
- 캐시 읽기/쓰기
- 종목코드 해석
- Restatement 반영
- 분기 코드 상수
- 네이버 wisereport 크롤링
- 엑셀 서식 적용
"""
import os
import json
import re
import pandas as pd
import requests
from bs4 import BeautifulSoup

# ── 캐시 ──
CACHE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'cache')
os.makedirs(CACHE_DIR, exist_ok=True)


def get_cache(cache_key, max_age_hours=None):
    """파일 캐시에서 데이터 로드. max_age_hours 지정 시 파일 수정 시간 기준 만료 체크."""
    cache_file = os.path.join(CACHE_DIR, f"{cache_key}.json")
    if os.path.exists(cache_file):
        if max_age_hours is not None:
            import time
            age_sec = time.time() - os.path.getmtime(cache_file)
            if age_sec > max_age_hours * 3600:
                return None
        with open(cache_file, 'r', encoding='utf-8') as f:
            return json.load(f)
    return None


def set_cache(cache_key, data):
    """파일 캐시에 데이터 저장"""
    cache_file = os.path.join(CACHE_DIR, f"{cache_key}.json")
    with open(cache_file, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False)


# ── 분기 코드 ──
QUARTER_CODES = {
    'Q1': ('11013', '1분기보고서'),
    'Q2': ('11012', '반기보고서'),
    'Q3': ('11014', '3분기보고서'),
    'Q4': ('11011', '사업보고서'),
}


# ── 종목코드 해석 ──
def resolve_stock_code(company, listed_companies):
    """종목코드/이름 → (stock_code, corp_name). 못 찾으면 (None, company)."""
    if company.isdigit() and len(company) == 6:
        stock_code = company
        match = listed_companies[listed_companies['stock_code'] == stock_code]
        corp_name = match.iloc[0]['corp_name'] if len(match) > 0 else company
        return stock_code, corp_name
    else:
        match = listed_companies[listed_companies['corp_name'] == company]
        if len(match) > 0:
            return match.iloc[0]['stock_code'], company
        return None, company


# ── Restatement 반영 ──
def apply_restatement(raw_data, years):
    """최신 사업보고서의 비교기간 데이터로 과거 연간값 재작성(restatement) 반영."""
    patched = set()
    for year in sorted(years, reverse=True):
        q4 = raw_data.get((year, 'Q4'))
        if q4 is None:
            continue
        for offset, col in [(1, 'frmtrm_amount'), (2, 'bfefrmtrm_amount')]:
            ty = year - offset
            if ty in patched or (ty, 'Q4') not in raw_data:
                continue
            if col not in q4.columns:
                continue
            old_df = raw_data[(ty, 'Q4')].copy()
            has_aid = 'account_id' in q4.columns and 'account_id' in old_df.columns
            has_sj = 'sj_nm' in q4.columns and 'sj_nm' in old_df.columns
            if has_aid:
                # (sj_nm, account_id) 복합키로 매핑하여 자본변동표 등
                # 다른 보고서의 동일 account_id가 덮어쓰는 문제 방지
                if has_sj:
                    rmap = {(str(r['sj_nm']), str(r['account_id'])): r[col]
                            for _, r in q4.iterrows()
                            if pd.notna(r.get(col))}
                    for idx, row in old_df.iterrows():
                        key = (str(row.get('sj_nm', '')), str(row.get('account_id', '')))
                        if key in rmap:
                            old_df.at[idx, 'thstrm_amount'] = rmap[key]
                else:
                    rmap = {str(r['account_id']): r[col] for _, r in q4.iterrows()
                            if pd.notna(r.get(col))}
                    for idx, row in old_df.iterrows():
                        aid = str(row.get('account_id', ''))
                        if aid in rmap:
                            old_df.at[idx, 'thstrm_amount'] = rmap[aid]
                raw_data[(ty, 'Q4')] = old_df
            patched.add(ty)


# ── 네이버 wisereport 크롤링 ──
NAVER_HEADERS = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'}


def crawl_wisereport(stock_code):
    """네이버 wisereport에서 기업정보 크롤링. 모든 가용 필드 반환."""
    result = {
        'description': '',
        'sector': '', 'sub_sector': '',
        'current_price': None, 'price_change': None, 'price_change_pct': None,
        'market_cap': None,
        'eps': None, 'bps': None, 'per': None, 'pbr': None,
        'div_yield': None, 'dps': None,
        'week52_high': None, 'week52_low': None,
        'shares_outstanding': None,
    }
    try:
        url = f'https://navercomp.wisereport.co.kr/v2/company/c1010001.aspx?cmp_cd={stock_code}'
        resp = requests.get(url, headers=NAVER_HEADERS, timeout=10)
        resp.encoding = 'utf-8'
        soup = BeautifulSoup(resp.text, 'html.parser')

        # 기업설명
        summary = soup.select_one('.cmp_comment')
        if summary:
            result['description'] = summary.get_text(strip=True)

        # dt 요소에서 업종/지표 추출
        for dt in soup.find_all('dt'):
            text = dt.get_text(strip=True)
            if (text.startswith('KOSPI') or text.startswith('KOSDAQ')) and ':' in text:
                sector_text = text.split(':', 1)[1].strip()
                for prefix in ['코스피', '코스닥', 'KOSPI', 'KOSDAQ']:
                    sector_text = sector_text.replace(prefix, '').strip()
                result['sector'] = sector_text
            elif text.startswith('WICS') and ':' in text:
                result['sub_sector'] = text.split(':', 1)[1].strip()
            elif text.startswith('EPS'):
                try: result['eps'] = float(text[3:].replace(',', '').strip())
                except: pass
            elif text.startswith('BPS'):
                try: result['bps'] = float(text[3:].replace(',', '').strip())
                except: pass
            elif text.startswith('PER') and '5년' not in text and '*' not in text:
                try: result['per'] = float(text[3:].replace(',', '').strip())
                except: pass
            elif text.startswith('PBR') and '*' not in text:
                try: result['pbr'] = float(text[3:].replace(',', '').strip())
                except: pass
            elif '배당수익률' in text and '*' not in text:
                try: result['div_yield'] = float(text.replace('배당수익률', '').replace('%', '').replace(',', '').strip())
                except: pass

        # 테이블에서 주가/시가총액/52주/배당 데이터
        for table in soup.find_all('table'):
            for tr in table.find_all('tr'):
                th = tr.find('th')
                td = tr.find('td')
                if not (th and td):
                    continue
                label = th.get_text(strip=True)
                val_text = td.get_text(strip=True)

                if '주가' in label and '전일대비' in label:
                    parts = val_text.split('/')
                    if parts:
                        try: result['current_price'] = int(parts[0].replace('원', '').replace(',', '').replace(' ', '').strip())
                        except: pass
                    if len(parts) >= 2:
                        try: result['price_change'] = int(parts[1].replace('원', '').replace(',', '').replace(' ', '').strip())
                        except: pass
                    if len(parts) >= 3:
                        try: result['price_change_pct'] = float(parts[2].replace('%', '').replace(' ', '').strip())
                        except: pass
                elif '시가총액' in label:
                    try: result['market_cap'] = int(val_text.replace('억', '').replace('원', '').replace(',', '').strip())
                    except: pass
                elif '52Weeks' in label or ('52주' in label and '베타' not in label):
                    nums = re.findall(r'[\d,]+', val_text)
                    if nums:
                        vals = [int(n.replace(',', '')) for n in nums if n.replace(',', '').isdigit()]
                        if len(vals) >= 2:
                            result['week52_high'] = max(vals)
                            result['week52_low'] = min(vals)
                elif '발행주식수' in label:
                    parts = val_text.split('/')
                    if parts:
                        try: result['shares_outstanding'] = int(parts[0].replace('주', '').replace(',', '').replace(' ', '').strip())
                        except: pass
                elif label == '현금DPS':
                    try: result['dps'] = int(float(val_text.replace('원', '').replace(',', '').strip()))
                    except: pass
                elif label == '배당수익률':
                    try: result['div_yield'] = float(val_text.replace('%', '').strip())
                    except: pass

    except Exception:
        pass
    return result


# ── 엑셀 서식 ──
def format_excel(excel_path, available_cols, config):
    """엑셀 파일에 표준 서식 적용.

    config = {
        'header_color': 'FFA500',       # 기본 헤더색
        'accent_color': '228B22',       # 강조 헤더색
        'accent_cols': [...],           # 강조 대상 컬럼명
        'decimal_cols': [...],          # 소수점 2자리 컬럼명
        'int_cols': [...],              # 정수(#,##0) 컬럼명
        'width_map': {'col': 15, ...},  # 컬럼별 너비 (없으면 default_width)
        'default_width': 10,
    }
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

        wb = load_workbook(excel_path)
        ws = wb.active

        header_fill = PatternFill(start_color=config['header_color'], end_color=config['header_color'], fill_type='solid')
        accent_fill = PatternFill(start_color=config['accent_color'], end_color=config['accent_color'], fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=10)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        accent_cols = set(config.get('accent_cols', []))
        decimal_indices = {i+1 for i, c in enumerate(available_cols) if c in config.get('decimal_cols', [])}
        int_indices = {i+1 for i, c in enumerate(available_cols) if c in config.get('int_cols', [])}
        width_map = config.get('width_map', {})
        default_width = config.get('default_width', 10)

        # 헤더
        for col_idx, cell in enumerate(ws[1], 1):
            col_name = available_cols[col_idx-1] if col_idx <= len(available_cols) else ''
            cell.fill = accent_fill if col_name in accent_cols else header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

        # 데이터
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for col_idx, cell in enumerate(row, 1):
                cell.border = thin_border
                if cell.value is not None:
                    if col_idx in decimal_indices:
                        cell.number_format = '0.00'
                        cell.alignment = Alignment(horizontal='right')
                    elif col_idx in int_indices:
                        cell.number_format = '#,##0'
                        cell.alignment = Alignment(horizontal='right')

        # 컬럼 너비
        for col_idx, col_name in enumerate(available_cols, 1):
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            ws.column_dimensions[col_letter].width = width_map.get(col_name, default_width)

        # 필터 + 틀고정
        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = 'A2'

        wb.save(excel_path)
    except Exception:
        pass
